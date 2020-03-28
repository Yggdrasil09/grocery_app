import pandas as pd
from flask import Flask, request, jsonify, json
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

def append_df_to_excel(filename, df):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    """
    from openpyxl import load_workbook

    import pandas as pd

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # try to open an existing workbook
    writer.book = load_workbook(filename)
    sheet_name = 'Sheet1'
    startrow = writer.book[sheet_name].max_row

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=None)

    # save the workbook
    writer.save()


def generate_order_number(filepath='data/ORDERS.xlsx'):
    df = pd.read_excel(filepath)
    alph = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
            'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    temp = str(100000 + len(df['ORDER ID']) + np.random.randint(1, 200000) + np.random.randint(1, 200000))
    return (np.random.choice(alph) + np.random.choice(alph) + temp)

app = Flask(__name__)

#food = pd.read_excel('data/inventory.xlsx', index_col='Item Code', sheet_name='Food')
#toiletry = pd.read_excel('data/inventory.xlsx', index_col='Item Code', sheet_name='Toiletry')
#cart = pd.DataFrame(columns=['Item', 'Quantity', 'Price'])
dropzone = pd.read_excel('data/dropzone.xlsx',usecols=['Drop Point', 'Drop Time'])
inventory = pd.read_excel('data/inventory_combined.xlsx', index_col='Item Code')


@app.route('/test',methods=['POST'])
def test():
    '''
    display = []
    for i in inventory['Item Name']:
        j = min(inventory[inventory['Item Name'] == i]['Item Qty'].values[0],
                inventory[inventory['Item Name'] == i]['Tab'].values[0])
        if j == 0:
            j = 'Item Out of Stock'
        k = inventory[inventory['Item Name'] == i]['Category'].values[0]
        l = inventory[inventory['Item Name'] == i]['Item Price'].values[0]
        display.append({'Item Name': i, 'Tab': str(j), 'Category':k, 'Price':str(l)})
    '''
    dic = {}
    for i in inventory['Category'].unique():
        dic[i] = []
    for key in dic.keys():
        temp = inventory[inventory['Category'] == key][['Item Name', 'Item Price', 'Tab', 'Item Qty']]
        for i in temp['Item Name']:
            j = min(temp[temp['Item Name'] == i]['Item Qty'].values[0],
                    temp[temp['Item Name'] == i]['Tab'].values[0])
            if j == 0:
                j = 'Item Out of Stock'
            l = temp[temp['Item Name'] == i]['Item Price'].values[0]
            dic[key].append({'Item Name': i, 'Tab': str(j), 'Price': str(l)})
    return jsonify(dic)




@app.route('/submit-order',methods=['GET','POST'])
def order_submit():

    data = {"Name": "Karan", "Zone": "CANT A",
            "Time": "11:30 AM",
            "Order": [{'item': 'Sugar', 'quantity': '3', 'category': 'food', 'price': '20'},
                      {'item': 'Toilet Paper', 'quantity': '1', 'category': 'toiletry', 'price': '30'},
                      {'item': 'Salt', 'quantity': '1', 'category': 'food', 'price': '10'},
                      {'item': 'Potatoes', 'quantity': '5', 'category': 'food', 'price': '25'}]}


    '''
    need to do this
    '''
    order_number = generate_order_number()

    df2 = pd.DataFrame(data['Order'])
    tc = ((df2['price'].apply(int)) * (df2['quantity'].apply(int))).sum()


    df1 = pd.DataFrame({'Order Number': str(order_number),
                        'Name': data['Name'],
                        'Zone': data['Zone'],
                        'Time': data['Time'],
                        'Total Cost': str(tc)}, index=[1])

    order = {'ORDER ID': str(order_number), 'ORDER': [data['Order']],
             'NAME': data['Name'], 'ZONE': data['Zone'], 'TIME': data['Time'], 'TOTAL COST': tc,
             'STATUS': 'Received'}

    df3 = pd.DataFrame(order)

    append_df_to_excel('data/ORDERS.xlsx',df3)

    for i in df2[['item', 'quantity']].iterrows():
        temp = inventory['Item Name'] == i[1].values[0]
        qty = inventory[temp]['Item Qty']
        inventory.loc[(temp), 'Item Qty'] = int(qty) - int(i[1].values[1])
        inventory.to_excel('data/inventory_combined.xlsx')

    fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1, figsize=(12, 4))

    ax1.axis('tight')
    ax1.axis('off')
    ax1.set_title('Bill')
    the_table1 = ax1.table(cellText=df1.values, colLabels=df1.columns, loc='center')
    ax2.set_title("Breakdown")
    ax2.axis('tight')
    ax2.axis('off')
    the_table2 = ax2.table(cellText=df2.values, colLabels=df2.columns, loc='center')
    pp = PdfPages("orders/Order Number {}.pdf".format(order_number))
    pp.savefig(fig, bbox_inches='tight')
    pp.close()

    return "Success"



if __name__ == '__main__':
    app.run(debug=True)